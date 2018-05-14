#!/usr/bin/env python2.7
# coding: utf8
"""Common function to be used by pws-booking.

        Most notably a function that converts a
        modern Excel spreadsheet (ie. an xlsx-formated
        spreadsheet) to a python dictionary.
        This *requires* the spreadsheet to have a header
        row that describes the contents of each column.
"""

import openpyxl

def xl_read_as_dict(spreadsheet):
    """Return a the first worksheet in a workbook as a dictionary"""

    book = openpyxl.reader.excel.load_workbook(spreadsheet)
    sheet = book.active

    rows = sheet.max_row
    cols = sheet.max_column

    # headers = dict((col, sheet.cell(row=1, column=col).value) for col in range(1, cols))

    def item(row, col):
        """Describe item key for column, ie. header."""
        return (sheet.cell(row=1, column=col).value, sheet.cell(row=row, column=col).value)

    return (dict(item(row, col) for col in range(1, cols + 1)) for row in range(2, rows + 1))
